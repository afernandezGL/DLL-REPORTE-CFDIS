"""Excel export helpers for writing the consolidated report and summary sheets."""

import logging
from copy import copy
import pandas as pd
from pathlib import Path
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from config.config import OUTPUT_FOLDER, EDICOM_LOG_FOLDER_NAME
from scr.models import (
    MONTH_MAP,
    DifferencesReportResult,
    SummaryResult,
    edicom_log_column_names,
    columns_to_export,
)
from scr.stryles import (
    BEIGE,
    BLACK,
    BLUE_BRIGHT,
    COLUMN_COLORS,
    COLUMN_GROUPS,
    COLUMN_GROUPS,
    DIFF_COLOR_NAMES,
    GREY,
    GROUP_COLORS,
    RED_BRIGHT,
    WHITE,
    YELLOW,
    diff_display_names,
)

logger = logging.getLogger(__name__)


def setup_resumen_header(
    ws: Workbook.active,
    section: str,
    company_name: str,
    report_title: str,
    year: str,
    month: str,
):
    ws["A2"] = section
    ws["A2"].font = Font(
        bold=True,
        color=RED_BRIGHT,
        size=14,
    )

    ws["A3"] = company_name
    ws["A3"].font = Font(
        bold=True,
        size=14,
    )

    ws.column_dimensions["A"].width = len(company_name) + 10

    ws["A4"] = report_title
    ws["A4"].font = Font(
        bold=True,
        size=14,
    )

    if month == "01":
        periodo = f"PERIODO: Enero {year}"
    else:
        periodo = f"PERIODO: Enero - {MONTH_MAP[int(month)]} {year}"

    ws["A5"] = periodo
    ws["A5"].font = Font(
        bold=True,
        size=14,
    )

    return 8


def add_metadata_detail_table(
    ws: Worksheet,
    metadata_df: pd.DataFrame,
    title_row: int = 8,
    title: str = "DETALLE METADATA",
):
    subtitle_row = title_row + 1
    header_row = subtitle_row + 1

    total_cols = len(metadata_df.columns)

    for col in range(1, total_cols + 1):
        ws.cell(title_row, col).fill = PatternFill(
            "solid",
            fgColor=GREY,
        )

    cell = ws.cell(title_row, 1)
    cell.value = title
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for col in range(1, total_cols + 1):
        ws.cell(subtitle_row, col).fill = PatternFill(
            "solid",
            fgColor=BLUE_BRIGHT,
        )

    cell = ws.cell(subtitle_row, 1)
    cell.value = "METADATA SAT"
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    for row_idx, row in enumerate(
        dataframe_to_rows(
            metadata_df,
            index=False,
            header=True,
        ),
        start=header_row,
    ):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=value,
            )

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(header_row, col_idx)

        cell.fill = PatternFill(
            "solid",
            fgColor=BLUE_BRIGHT,
        )

        cell.font = Font(
            bold=True,
            color=WHITE,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.auto_filter.ref = (
        f"A{header_row}:" f"{ws.cell(ws.max_row, ws.max_column).coordinate}"
    )


def add_edicom_detail_table(
    ws: Worksheet,
    edicom_df: pd.DataFrame,
    title_row: int = 8,
    title: str = "EDICOM",
):
    subtitle_row = title_row + 1
    header_row = subtitle_row + 1

    total_cols = len(edicom_df.columns)

    for col in range(1, total_cols + 1):
        ws.cell(title_row, col).fill = PatternFill(
            "solid",
            fgColor=GREY,
        )

    cell = ws.cell(title_row, 1)
    cell.value = title
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for col in range(1, total_cols + 1):
        ws.cell(subtitle_row, col).fill = PatternFill(
            "solid",
            fgColor=BEIGE,
        )

    cell = ws.cell(subtitle_row, 1)
    cell.value = "DETALLE EDICOM"
    cell.font = Font(bold=True, color=BLACK)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    for row_idx, row in enumerate(
        dataframe_to_rows(
            edicom_df,
            index=False,
            header=True,
        ),
        start=header_row,
    ):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=value,
            )

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(header_row, col_idx)

        cell.fill = PatternFill(
            "solid",
            fgColor=BEIGE,  # o el color correspondiente
        )

        cell.font = Font(
            bold=True,
            color=BLACK,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.auto_filter.ref = (
        f"A{header_row}:" f"{ws.cell(ws.max_row, ws.max_column).coordinate}"
    )


def add_cfdi_detail_table(
    ws: Worksheet,
    cfdi_df: pd.DataFrame,
    title_row: int = 8,
    title: str = "DETALLE CFDI",
):
    subtitle_row = title_row + 1
    header_row = subtitle_row + 1

    total_cols = len(cfdi_df.columns)

    for col in range(1, total_cols + 1):
        ws.cell(title_row, col).fill = PatternFill(
            "solid",
            fgColor=GREY,
        )

    cell = ws.cell(title_row, 1)
    cell.value = title
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for col in range(1, total_cols + 1):
        ws.cell(subtitle_row, col).fill = PatternFill(
            "solid",
            fgColor=RED_BRIGHT,
        )

    cell = ws.cell(subtitle_row, 1)
    cell.value = "DETALLE CFDI"
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for row_idx, row in enumerate(
        dataframe_to_rows(
            cfdi_df,
            index=False,
            header=True,
        ),
        start=header_row,
    ):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=value,
            )

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(header_row, col_idx)

        cell.fill = PatternFill(
            "solid",
            fgColor=RED_BRIGHT,
        )

        cell.font = Font(
            bold=True,
            color=WHITE,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.auto_filter.ref = (
        f"A{header_row}:" f"{ws.cell(ws.max_row, ws.max_column).coordinate}"
    )


def add_consolidated_detail_table(
    cols_present: list,
    ws: Worksheet,
    consolidated_df: pd.DataFrame,
    title_row: int = 8,
    title: str = "CONSOLIDADO",
):
    subtitle_row = title_row + 1
    header_row = subtitle_row + 1

    cell = ws.cell(title_row, 1)

    ws.cell(title_row, 1).value = title

    cell.fill = PatternFill("solid", fgColor=GREY)

    cell.font = Font(bold=True, color=WHITE)

    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(cols_present) + 1):
        ws.cell(title_row, col).fill = PatternFill(
            fill_type="solid",
            fgColor=GREY,
        )

    start_col = 1
    current_group = None

    for idx, col_name in enumerate(columns_to_export, start=1):
        group = COLUMN_GROUPS.get(col_name)

        if current_group is None:
            current_group = group
            start_col = idx

        elif group != current_group:
            end_col = idx - 1
            group_color = GROUP_COLORS.get(current_group, "D9D9D9")

            # Pintar todas las columnas del grupo
            for col in range(start_col, end_col + 1):
                ws.cell(subtitle_row, col).fill = PatternFill(
                    fill_type="solid",
                    fgColor=group_color,
                )

            # SIEMPRE escribir el texto del grupo
            cell = ws.cell(subtitle_row, start_col)
            cell.value = current_group

            cell.font = Font(
                bold=True,
                color=(WHITE if group_color in [RED_BRIGHT, BLUE_BRIGHT] else BLACK),
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            # Preparar siguiente grupo
            current_group = group
            start_col = idx

    # ==========================================
    # Último grupo
    # ==========================================

    end_col = len(columns_to_export)

    group_color = GROUP_COLORS.get(current_group, "D9D9D9")

    for col in range(start_col, end_col + 1):
        ws.cell(subtitle_row, col).fill = PatternFill(
            fill_type="solid",
            fgColor=group_color,
        )

    cell = ws.cell(subtitle_row, start_col)
    cell.value = current_group

    cell.font = Font(
        bold=True,
        color=(WHITE if group_color in [RED_BRIGHT, BLUE_BRIGHT] else BLACK),
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    for row in dataframe_to_rows(
        consolidated_df[cols_present], index=False, header=True
    ):
        ws.append(row)

    thin = Side(style="thin", color=BLACK)

    for cell in ws[header_row]:
        color = COLUMN_COLORS.get(str(cell.value).strip())

        if color:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=color,
            )

        cell.font = Font(
            bold=True,
            color=WHITE if color in (BLUE_BRIGHT, RED_BRIGHT) else BLACK,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

    for column in ws.columns:
        max_length = 0

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(column[0].column)].width = min(
            max_length + 3, 40
        )

    ws.auto_filter.ref = (
        f"A{header_row}:" f"{ws.cell(ws.max_row, ws.max_column).coordinate}"
    )

    headers = [cell.value for cell in ws[header_row]]
    money_columns = [
        "SUBTOTAL",
        "IVA",
        "TOTAL",
        "TOTAL CONCEPTO",
        "Tipo de cambio",
        "TOTAL METADATA",
        "TOTAL CONCEPTO MXN",
    ]
    for col_name in money_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "#,##0.00"


def add_subtotal_differences_detail_table(
    ws: Worksheet,
    subtotal_differences_df: pd.DataFrame,
    title_row: int = 8,
    title: str = "DIFERENCIAS POR SUBTOTALES",
):
    total_cols = len(subtotal_differences_df.columns)

    # ======================
    # TÍTULO
    # ======================
    for col in range(1, total_cols + 1):
        ws.cell(title_row, col).fill = PatternFill(
            "solid",
            fgColor=GREY,
        )

    cell = ws.cell(title_row, 1)
    cell.value = title
    cell.font = Font(
        bold=True,
        color=WHITE,
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    # ======================
    # SUBTÍTULO
    # ======================
    subtitle_row = title_row + 1

    for col in range(1, total_cols + 1):
        ws.cell(subtitle_row, col).fill = PatternFill(
            "solid",
            fgColor=YELLOW,
        )

    cell = ws.cell(subtitle_row, 1)
    cell.value = "COMPARATIVO"
    cell.font = Font(
        bold=True,
        color=BLACK,
    )
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    # ======================
    # ENCABEZADOS
    # ======================
    header_row = subtitle_row + 1

    for col_idx, real_col in enumerate(
        subtotal_differences_df.columns,
        start=1,
    ):
        header_cell = ws.cell(
            row=header_row,
            column=col_idx,
        )

        header_cell.value = diff_display_names.get(
            real_col,
            str(real_col),
        )

        color = DIFF_COLOR_NAMES.get(
            real_col,
            YELLOW,
        )

        header_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=color,
        )

        header_cell.font = Font(
            bold=True,
            color=WHITE if color in (BLUE_BRIGHT, RED_BRIGHT, GREY) else BLACK,
        )
    # ======================
    # DATOS
    # ======================
    for row_idx, (_, row) in enumerate(
        subtotal_differences_df.iterrows(),
        start=header_row + 1,
    ):
        for col_idx, real_col in enumerate(
            subtotal_differences_df.columns,
            start=1,
        ):
            ws.cell(
                row=row_idx,
                column=col_idx,
                value=row[real_col],
            )

    current_row = header_row + len(subtotal_differences_df) + 1

    return current_row


def export_to_winba_format(
    consolidated_df: pd.DataFrame,
    raw_metadata_info_df: pd.DataFrame,
    summaries: SummaryResult,
    differences: DifferencesReportResult,
    date_: str,
) -> bool:
    """Export the consolidated dataset and monthly summaries to an Excel workbook.

    The function creates a workbook with a detailed output sheet and a summary
    sheet containing the Edicom, metadata, and CFDI recap sections.

    Args:
        consolidated_df: The fully processed and reconciled dataset.
        summaries: Summary statistics for Edicom, metadata, and CFDI invoice results.
        differences: DifferencesReportResult containing UUID and subtotal differences.
        date_: Period identifier used in the output file name.

    Returns:
        True when the export completes successfully.
    """

    output_dir = OUTPUT_FOLDER / date_
    year, month = date_.split("_")
    output_dir.mkdir(parents=True, exist_ok=True)
    file_name = f"Sofom - Amarre de facturación Invoicing con MTD SAT {date_}.xlsx"
    output_path = output_dir / file_name

    missing = [c for c in columns_to_export if c not in consolidated_df.columns]
    if missing:
        logger.warning(
            "Some columns to export are missing", extra={"missing_columns": missing}
        )

    cols_present = [c for c in columns_to_export if c in consolidated_df.columns]
    # consolidated_df[cols_present].to_excel(output_path, index=False)

    # Crear workbook
    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "RESUMEN"

    setup_resumen_header(
        wb.active,
        "A.1",
        "DLL LEASING / DE LAGE LANDEN",
        "Resumen Conciliación",
        year,
        month,
    )

    metadata_title_row = add_winba_resumen_block(
        ws=ws_resumen,
        current_row=8,
        title="EDICOM",
        df=summaries.edicom,
        color=COLUMN_COLORS["ESTATUS_EDICOM"],
    )
    factura_title_row = add_winba_resumen_block(
        ws=ws_resumen,
        current_row=metadata_title_row,
        title="METADATA",
        df=summaries.metadata,
        color=COLUMN_COLORS["ESTATUS"],
    )
    add_winba_resumen_block(
        ws=ws_resumen,
        current_row=factura_title_row,
        title="FACTURA",
        df=summaries.factura,
        color=COLUMN_COLORS["USO CFDI"],
    )
    autofit_columns(ws_resumen)

    ws_consolidated = wb.create_sheet("CONSOLIDADO")
    setup_resumen_header(
        ws_consolidated,
        "A.1.1",
        "DLL LEASING / DE LAGE LANDEN",
        "Conciliación Edicom vs Metadata vs Facturas",
        year,
        month,
    )
    add_consolidated_detail_table(
        cols_present,
        ws_consolidated,
        consolidated_df,
        title_row=8,
        title="CONCILIACIÓN EDICOM vs METADATA vs PARSEO FACTURAS",
    )
    autofit_columns(ws_consolidated, ignore_columns=["S"])

    ws_metadata = wb.create_sheet("METADATA")
    setup_resumen_header(
        ws_metadata,
        "A.1.2",
        "DLL LEASING / DE LAGE LANDEN",
        "Detalle Metadata",
        year,
        month,
    )
    add_metadata_detail_table(
        ws_metadata, raw_metadata_info_df, title_row=8, title="DETALLE METADATA"
    )
    autofit_columns(ws_metadata)

    ws_diferencias = wb.create_sheet("DIFERENCIAS CONSOLIDADO")
    setup_resumen_header(
        ws_diferencias,
        "A.1.3",
        "DLL LEASING / DE LAGE LANDEN",
        (
            "Detalle Diferencias por UUID"
            if differences.consolidated.shape[0] > 0
            else "No hay diferencias de UUID"
        ),
        year,
        month,
    )
    add_differences_table(
        ws_diferencias,
        differences.consolidated,
        title_row=8,
        title="DIFERENCIAS DE UUID",
    )
    add_metadata_detail_table(
        ws_diferencias,
        differences.uuid.metadata,
        title_row=ws_diferencias.max_row + 2,
        title="METADATA DIFERENCIAS",
    )
    add_cfdi_detail_table(
        ws_diferencias,
        differences.uuid.factura,
        title_row=ws_diferencias.max_row + 2,
        title="FACTURAS DIFERENCIAS",
    )
    add_edicom_detail_table(
        ws_diferencias,
        differences.uuid.edicom,
        title_row=ws_diferencias.max_row + 2,
        title="EDICOM DIFERENCIAS",
    )
    autofit_columns(ws_diferencias)

    ws_diferencias_subtotal = wb.create_sheet("SUBTOTALES")
    setup_resumen_header(
        ws_diferencias_subtotal,
        "A.1.4",
        "DLL LEASING / DE LAGE LANDEN",
        "Detalle subtotales",
        year,
        month,
    )
    add_subtotal_differences_detail_table(
        ws_diferencias_subtotal,
        differences.comparative_subtotals,
        title_row=8,
        title="DIFERENCIAS DE SUBTOTALES",
    )
    autofit_columns(ws_diferencias_subtotal)

    ws_diferencias_subtotal = wb.create_sheet("DIFERENCIAS SUBTOTAL")
    setup_resumen_header(
        ws_diferencias_subtotal,
        "A.1.5",
        "DLL LEASING / DE LAGE LANDEN",
        (
            "Detalle Diferencias por subtotales"
            if differences.comparative_subtotals.shape[0] > 0
            else "No hay diferencias de subtotales"
        ),
        year,
        month,
    )
    add_subtotal_differences_detail_table(
        ws_diferencias_subtotal,
        differences.relevant_comparative_subtotals,
        title_row=8,
        title="DIFERENCIAS DE SUBTOTALES",
    )
    add_metadata_detail_table(
        ws_diferencias_subtotal,
        differences.uuid.metadata,
        title_row=ws_diferencias_subtotal.max_row + 2,
        title="METADATA DIFERENCIAS",
    )
    add_cfdi_detail_table(
        ws_diferencias_subtotal,
        differences.uuid.factura,
        title_row=ws_diferencias_subtotal.max_row + 2,
        title="FACTURAS DIFERENCIAS",
    )
    add_edicom_detail_table(
        ws_diferencias_subtotal,
        differences.uuid.edicom,
        title_row=ws_diferencias_subtotal.max_row + 2,
        title="EDICOM DIFERENCIAS",
    )
    autofit_columns(ws_diferencias_subtotal, ignore_columns=["T"])

    return save_file(wb, date_)


def export_to_client_format(
    consolidated_df: pd.DataFrame,
    edicom_resumen: pd.DataFrame,
    metadata_resumen: pd.DataFrame,
    factura_resumen: pd.DataFrame,
    uuid_differences: pd.DataFrame,
    differences_dict: dict,
    subtotal_differences: pd.DataFrame,
    date_: str,
) -> bool:
    """Export the consolidated dataset and monthly summaries to an Excel workbook.

    The function creates a workbook with a detailed output sheet and a summary
    sheet containing the Edicom, metadata, and CFDI recap sections.

    Args:
        consolidated_df: The fully processed and reconciled dataset.
        edicom_resumen: Summary statistics for Edicom results.
        metadata_resumen: Summary statistics for metadata results.
        factura_resumen: Summary statistics for CFDI invoice results.
        date_: Period identifier used in the output file name.

    Returns:
        True when the export completes successfully.
    """

    year = date_.split("_")[0]
    missing = [c for c in columns_to_export if c not in consolidated_df.columns]
    if missing:
        logger.warning(
            "Some columns to export are missing", extra={"missing_columns": missing}
        )

    cols_present = [c for c in columns_to_export if c in consolidated_df.columns]
    # consolidated_df[cols_present].to_excel(output_path, index=False)

    # Crear workbook
    wb = Workbook()

    # =====================================================
    # Sheet 1 - Salida
    # =====================================================

    ws_salida = wb.active
    ws_salida.title = "Salida"

    for row in dataframe_to_rows(
        consolidated_df[cols_present], index=False, header=True
    ):
        ws_salida.append(row)

    thin = Side(style="thin", color=BLACK)

    for cell in ws_salida[1]:
        color = COLUMN_COLORS.get(str(cell.value).strip())

        if color:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=color,
            )

        cell.font = Font(
            bold=True,
            color=WHITE if color == BLUE_BRIGHT else BLACK,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )

    # Filtros
    ws_salida.auto_filter.ref = ws_salida.dimensions

    # Congelar encabezado
    ws_salida.freeze_panes = "A2"

    # Altura encabezado
    ws_salida.row_dimensions[1].height = 45

    for column in ws_salida.columns:
        max_length = 0

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws_salida.column_dimensions[get_column_letter(column[0].column)].width = min(
            max_length + 3, 40
        )

    headers = [cell.value for cell in ws_salida[1]]
    money_columns = [
        "SUBTOTAL",
        "IVA",
        "TOTAL",
        "TOTAL CONCEPTO",
        "Tipo de cambio",
        "TOTAL METADATA",
        "TOTAL CONCEPTO MXN",
    ]
    for col_name in money_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            for row in range(2, ws_salida.max_row + 1):
                ws_salida.cell(row=row, column=col_idx).number_format = "#,##0.00"

    # =====================================================
    # Sheet 2 - Resumen
    # =====================================================

    ws_resumen = wb.create_sheet("Resumen")

    edicom_color = COLUMN_COLORS["ESTATUS_EDICOM"]
    metadata_color = COLUMN_COLORS["ESTATUS METADATA"]
    factura_color = COLUMN_COLORS["USO CFDI"]

    # Título principal
    ws_resumen["A1"] = year
    ws_resumen["A1"].font = Font(size=16, bold=True)

    current_row = 3

    # =====================================================
    # EDICOM
    # =====================================================

    edicom_title_row = current_row

    ws_resumen.cell(current_row, 1, "EDICOM")
    ws_resumen.cell(current_row, 1).font = Font(size=12, bold=True)

    header_row = edicom_title_row + 1
    current_row += 1

    for row in dataframe_to_rows(edicom_resumen, index=False, header=True):
        ws_resumen.append(row)

    # Título EDICOM
    ws_resumen.cell(edicom_title_row, 1).fill = PatternFill(
        fill_type="solid",
        fgColor=edicom_color,
    )
    ws_resumen.cell(edicom_title_row, 1).font = Font(
        bold=True,
        size=12,
    )

    # Encabezados de la tabla EDICOM (fila 4)
    for cell in ws_resumen[header_row]:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=edicom_color,
        )
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    current_row = ws_resumen.max_row + 2
    metadata_title_row = current_row

    # =====================================================
    # METADATA
    # =====================================================

    ws_resumen.cell(metadata_title_row, 1, "METADATA")
    ws_resumen.cell(metadata_title_row, 1).font = Font(size=12, bold=True)

    header_row = metadata_title_row + 1
    current_row += 1

    for row in dataframe_to_rows(metadata_resumen, index=False, header=True):
        ws_resumen.append(row)

    # Título METADATA
    ws_resumen.cell(metadata_title_row, 1).fill = PatternFill(
        fill_type="solid",
        fgColor=metadata_color,
    )
    ws_resumen.cell(metadata_title_row, 1).font = Font(
        bold=True,
    )

    # Encabezados de la tabla METADATA
    for cell in ws_resumen[header_row]:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=metadata_color,
        )

        cell.font = Font(bold=True)

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    current_row = ws_resumen.max_row + 2
    factura_title_row = current_row

    # =====================================================
    # FACTURA
    # =====================================================

    ws_resumen.cell(factura_title_row, 1, "FACTURA")
    ws_resumen.cell(factura_title_row, 1).font = Font(size=12, bold=True)

    header_row = factura_title_row + 1
    current_row += 1

    for row in dataframe_to_rows(factura_resumen, index=False, header=True):
        ws_resumen.append(row)

    # Título FACTURA
    ws_resumen.cell(factura_title_row, 1).fill = PatternFill(
        fill_type="solid",
        fgColor=factura_color,
    )
    ws_resumen.cell(factura_title_row, 1).font = Font(
        bold=True,
        size=12,
    )

    # Encabezados de la tabla FACTURA
    for cell in ws_resumen[header_row]:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=factura_color,
        )

        cell.font = Font(
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column in ws_resumen.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws_resumen.column_dimensions[column_letter].width = max_length + 3

    # Format numeric columns
    for row in ws_resumen.iter_rows(min_row=1):
        for cell in row:
            # B and D -> counts
            if cell.column in [3, 4, 5, 6]:
                cell.number_format = "#,##0"

    ws_diferencias = wb.create_sheet("Diferencias uuid")

    add_differences_table(
        ws_diferencias, uuid_differences, title_row=1, title="DIFERENCIAS DE UUID"
    )
    add_metadata_detail_table(
        ws_diferencias,
        differences_dict["UUID"]["metadata"],
        title_row=ws_diferencias.max_row + 2,
        title="EDICOM DIFERENCIAS",
    )
    add_cfdi_detail_table(
        ws_diferencias,
        differences_dict["UUID"]["facturas"],
        title_row=ws_diferencias.max_row + 2,
        title="CFDI DIFERENCIAS",
    )
    add_consolidated_detail_table(
        cols_present,
        ws_diferencias,
        differences_dict["UUID"]["edicom"],
        title_row=ws_diferencias.max_row + 2,
        title="EDICOM DIFERENCIAS",
    )
    autofit_columns(ws_diferencias)

    ws_diferencias_subtotal = wb.create_sheet("Diferencias subtotal")
    add_consolidated_detail_table(
        cols_present,
        ws_diferencias_subtotal,
        subtotal_differences,
        title_row=1,
        title="DIFERENCIAS DE SUBTOTALES",
    )
    autofit_columns(ws_diferencias_subtotal)

    return save_file(wb, date_)


def add_differences_table(
    ws_diferencias: Worksheet,
    uuid_differences: pd.DataFrame,
    title_row: int = 0,
    title: str = "DIFERENCIAS DE UUID",
):
    subtitle_row = title_row + 1
    header_row = subtitle_row + 1

    total_cols = len(uuid_differences.columns)

    for col in range(1, total_cols + 1):
        ws_diferencias.cell(title_row, col).fill = PatternFill(
            "solid",
            fgColor=GREY,
        )

    cell = ws_diferencias.cell(title_row, 1)
    cell.value = title
    cell.font = Font(bold=True, color=WHITE)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for col in range(1, total_cols + 1):
        ws_diferencias.cell(subtitle_row, col).fill = PatternFill(
            "solid",
            fgColor=YELLOW,
        )

    cell = ws_diferencias.cell(subtitle_row, 1)
    cell.value = "COMPARATIVO"
    cell.font = Font(bold=True, color=BLACK)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for row in dataframe_to_rows(uuid_differences, index=False, header=True):
        ws_diferencias.append(row)

    for cell in ws_diferencias[header_row]:
        cell.fill = PatternFill("solid", fgColor=YELLOW)

        cell.font = Font(bold=True, color=BLACK)

        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ws.freeze_panes = f"A{header_row + 1}"

    ws_diferencias.auto_filter.ref = (
        f"A{header_row}:"
        f"{ws_diferencias.cell(ws_diferencias.max_row, ws_diferencias.max_column).coordinate}"
    )


def save_file(wb: Workbook, date_: str) -> bool:
    try:
        output_dir = OUTPUT_FOLDER / date_
        output_dir.mkdir(parents=True, exist_ok=True)
        file_path = (
            output_dir
            / f"Sofom - Amarre de facturación Invoicing con MTD SAT {date_}.xlsx"
        )
        wb.save(file_path)
        logger.info("Exported consolidated dataframe to Excel")
    except Exception as e:
        logger.error(
            "Failed to save Excel file",
            extra={"output_path": str(output_dir), "error": str(e)},
        )
        return False
    return True


def save_log(normalize_transformed_edicom_info_df: pd.DataFrame, date_: str) -> bool:
    """Persist the transformed Edicom log rows to an Excel workbook.

    Args:
        normalize_transformed_edicom_info_df: The transformed Edicom rows to store.
        date_: Period identifier used to locate the target log file.

    Returns:
        True when the log workbook is written successfully.
    """
    year = date_.split("_")[0]
    month = date_.split("_")[1]
    month = MONTH_MAP[int(month)]
    output_dir = EDICOM_LOG_FOLDER_NAME / year
    file_name = f"Log_{month}.xlsx"
    output_path = output_dir / file_name

    normalize_transformed_edicom_info_df[edicom_log_column_names].to_excel(
        output_path, index=False
    )
    logger.info(
        "Exported log dataframe to Excel",
        extra={
            "output_path": str(output_path),
            "rows": int(normalize_transformed_edicom_info_df.shape[0]),
        },
    )
    return True


def add_winba_resumen_block(ws, current_row, title, df, color):
    """Add a summary block to the worksheet with a title and a table.

    Args:
        ws: The worksheet to modify.
        current_row: The row at which to start the block.
        title: The title of the summary block.
        df: The dataframe containing the summary data.
        color: The background color for the title and header.

    Returns:
        The row after the summary block.
    """
    title_row = current_row
    header_row = current_row + 1

    num_cols = len(df.columns)

    # Título
    ws.merge_cells(
        start_row=title_row, start_column=1, end_row=title_row, end_column=num_cols
    )

    title_cell = ws.cell(title_row, 1)
    title_cell.value = title
    title_cell.fill = PatternFill(fill_type="solid", fgColor=color)
    title_cell.font = Font(
        bold=True, color=WHITE if color == RED_BRIGHT else BLACK, size=12
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Tabla
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Encabezados
    for cell in ws[header_row]:
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.font = Font(bold=True, color=WHITE if color == RED_BRIGHT else BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    headers = [cell.value for cell in ws[header_row]]

    # Columnas de conteo
    count_columns = {
        "N_FACTURAS_VIGENTES",
        "N_FACTURAS_CANCELADAS",
    }

    # Columnas monetarias
    money_columns = {
        "TOTAL_VIGENTES",
        "TOTAL_CANCELADAS",
    }

    # Formato numérico
    for col_name in count_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1

            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "#,##0"

    # Formato moneda
    for col_name in money_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1

            for row in range(header_row + 1, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = "$* #,##0"
    # Fila TOTAL en negritas
    for row in range(header_row + 1, ws.max_row + 1):
        periodo_value = ws.cell(row, 2).value
        if str(periodo_value).strip().upper() == "TOTAL":
            for cell in ws[ws.max_row]:
                new_font = copy(cell.font)
                new_font.bold = True
                cell.font = new_font
            break

    return ws.max_row + 2


def add_client_resume_block(ws, current_row, title, df, color):
    title_row = current_row
    header_row = current_row + 1

    num_cols = len(df.columns)

    ws.merge_cells(
        start_row=title_row, start_column=1, end_row=title_row, end_column=num_cols
    )

    cell = ws.cell(title_row, 1)
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color=WHITE if color == RED_BRIGHT else BLACK, size=12)
    cell.alignment = Alignment(horizontal="center")

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=WHITE if color == RED_BRIGHT else BLACK)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    return ws.max_row + 2


def apply_number_format_table(ws_resumen):
    # Format numeric columns
    for row in ws_resumen.iter_rows(min_row=1):
        for cell in row:
            # B and D -> counts
            if cell.column in [3, 4, 5, 6]:
                cell.number_format = "#,##0"


def autofit_columns(ws, ignore_columns=None):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        if ignore_columns and column_letter in ignore_columns:
            continue

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3
