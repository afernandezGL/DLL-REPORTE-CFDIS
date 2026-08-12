"""Small unit tests for the executable modules in ``scr``."""

import sys
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import Workbook

from scr import database, export, integration, loader, orchestrator, report_pipeline
from scr.models import RawResult, ReportResult, TransformedResult
from scr.transformer import (
    filter_metadata_info,
    normalize_concepts,
    transform_cfdi_info,
)


def test_get_engine_rejects_missing_database_configuration(monkeypatch):
    """Database creation fails clearly when a required setting is absent."""
    monkeypatch.setattr(database, "DB_SERVER", None)

    with pytest.raises(ValueError):
        database.get_engine()


def test_get_engine_builds_engine_with_configured_values(monkeypatch):
    """Database creation passes the configured connection string to SQLAlchemy."""
    captured = {}

    def fake_create_engine(connection_string, pool_pre_ping):
        captured["connection_string"] = connection_string
        captured["pool_pre_ping"] = pool_pre_ping
        return "engine"

    monkeypatch.setattr(database, "DB_SERVER", "server")
    monkeypatch.setattr(database, "DB_DATABASE", "database")
    monkeypatch.setattr(database, "DB_USER", "user")
    monkeypatch.setattr(database, "DB_PASSWORD", "password")
    monkeypatch.setattr(database, "DB_PORT", "1433")
    monkeypatch.setattr(database, "create_engine", fake_create_engine)

    assert database.get_engine() == "engine"
    assert captured == {
        "connection_string": "mssql+pymssql://user:password@server:1433/database",
        "pool_pre_ping": True,
    }


def test_close_engine_disposes_resource():
    """Closing an engine calls its dispose method."""
    engine = SimpleNamespace(dispose=lambda: setattr(engine, "disposed", True))
    engine.disposed = False

    database.close_engine(engine)

    assert engine.disposed is True


def test_setup_resumen_header_writes_period_and_returns_start_row():
    """Summary headers contain the requested section and period."""
    workbook = Workbook()
    worksheet = workbook.active

    next_row = export.setup_resumen_header(
        worksheet, "A.1", "Company", "Report", "2026", "01"
    )

    assert next_row == 8
    assert worksheet["A2"].value == "A.1"
    assert worksheet["A5"].value == "PERIODO: Enero 2026"


def test_add_differences_table_writes_headers_and_rows():
    """UUID differences are written with a title, headers, and data rows."""
    workbook = Workbook()
    worksheet = workbook.active
    differences = pd.DataFrame(
        {"UUID": ["uuid-1"], "EDICOM": [True], "METADATA": [False]}
    )

    export.add_differences_table(worksheet, differences, title_row=1)

    assert worksheet["A1"].value == "DIFERENCIAS DE UUID"
    assert worksheet["A3"].value == "UUID"
    assert worksheet["A4"].value == "uuid-1"


def test_add_client_resume_block_returns_row_after_table():
    """Client summary blocks return a row after the rendered content."""
    workbook = Workbook()
    worksheet = workbook.active
    summary = pd.DataFrame({"PERIODO": ["TOTAL"], "TOTAL": [10]})

    next_row = export.add_client_resume_block(
        worksheet, 1, "RESUMEN", summary, "FFFF00"
    )

    assert worksheet["A1"].value == "RESUMEN"
    assert worksheet["A2"].value == "PERIODO"
    assert next_row == 5


def test_select_best_matches_keeps_unique_concept_and_invoice_ids():
    """Candidate selection does not reuse either side of a match."""
    candidates = pd.DataFrame(
        {
            "CONCEPTO_ID": [1, 1, 2],
            "CFDI_ID": [10, 11, 11],
            "value": ["first", "duplicate concept", "duplicate invoice"],
        }
    )

    result = integration.select_best_matches(candidates)

    assert result["value"].tolist() == ["first", "duplicate invoice"]
    assert result["CONCEPTO_ID"].is_unique
    assert result["CFDI_ID"].is_unique


def test_build_uuid_differences_flags_missing_sources():
    """UUID comparison identifies the source in which each UUID is present."""
    consolidated = pd.DataFrame({"UUID": ["u1", "u2"]})
    metadata = pd.DataFrame({"Uuid": ["u1", "u3"]})
    factura = pd.DataFrame({"UUID": ["u1", "u4"]})

    result = integration.build_uuid_differences(consolidated, metadata, factura)
    row = result.set_index("UUID")

    assert bool(row.loc["u2", "EDICOM"]) is True
    assert bool(row.loc["u2", "METADATA"]) is False
    assert bool(row.loc["u4", "FACTURAS"]) is True


def test_system_presence_helpers_describe_comparison_row():
    """Presence helpers return the expected comma-separated system names."""
    row = pd.Series({"FACTURAS": True, "EDICOM": False, "METADATA": True})

    assert integration.get_missing_systems(row) == "EDICOM"
    assert integration.get_present_systems(row) == "FACTURAS, METADATA"


def test_get_metadata_info_reads_zip_text_file(tmp_path, monkeypatch):
    """Metadata loading combines a delimited file from a ZIP archive."""
    import zipfile

    period_folder = tmp_path / "2026_01"
    period_folder.mkdir()
    zip_path = period_folder / "metadata.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.writestr("metadata.txt", "RfcEmisor~Uuid\nAAA~u1\n")
    monkeypatch.setattr(loader, "METADATA_FOLDER_NAME", tmp_path)

    result = loader.get_metadata_info("2026_01")

    assert result.shape == (1, 2)
    assert result.loc[0, "Uuid"] == "u1"


def test_get_edicom_logs_returns_expected_empty_january_structure():
    """January starts the annual Edicom log with the expected columns."""
    result = loader.get_edicom_logs("2026_01")

    assert result.empty
    assert result.columns.tolist() == loader.edicom_log_column_names


def test_get_cfdi_info_closes_engine_after_query(monkeypatch):
    """CFDI loading closes its engine after the query completes."""
    closed = []
    monkeypatch.setattr(loader, "get_engine", lambda: "engine")
    monkeypatch.setattr(loader, "close_engine", lambda engine: closed.append(engine))
    monkeypatch.setattr(
        loader.pd,
        "read_sql",
        lambda query, engine: pd.DataFrame({"UUID": ["u1"]}),
    )

    result = loader.get_cfdi_info("2026_01", ["AAA"])

    assert result["UUID"].tolist() == ["u1"]
    assert closed == ["engine"]


def test_report_result_stores_three_source_dataframes():
    """ReportResult preserves the three source dataframes."""
    frames = [pd.DataFrame({"value": [1]}) for _ in range(3)]

    result = ReportResult(*frames)

    assert result.edicom is frames[0]
    assert result.metadata is frames[1]
    assert result.factura is frames[2]


def test_raw_result_is_a_report_result():
    """RawResult inherits the common report container contract."""
    result = RawResult(pd.DataFrame(), pd.DataFrame(), pd.DataFrame())

    assert isinstance(result, ReportResult)


def test_transformed_result_keeps_extra_transformation_frames():
    """TransformedResult stores filtered metadata and normalized Edicom data."""
    result = TransformedResult(
        pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    )

    assert result.filtered_metadata.empty
    assert result.normalize_edicom.empty


def test_filter_by_uuid_uses_requested_column():
    """UUID filtering keeps only rows with requested identifiers."""
    frame = pd.DataFrame({"id": ["u1", "u2"], "value": [1, 2]})

    result = orchestrator.filter_by_uuid(frame, ["u2"], uuid_column="id")

    assert result["value"].tolist() == [2]


def test_consolidate_info_delegates_to_join_and_integrate(monkeypatch):
    """Consolidation invokes the join and integration stages in order."""
    calls = []
    joined = pd.DataFrame({"UUID": ["u1"]})
    integrated = pd.DataFrame({"UUID": ["u1"], "PREFIJO": ["REN"]})
    monkeypatch.setattr(
        orchestrator, "join_dfs", lambda result: calls.append("join") or joined
    )
    monkeypatch.setattr(
        orchestrator,
        "integrate_data",
        lambda frame: calls.append("integrate") or integrated,
    )

    result = orchestrator.consolidate_info(SimpleNamespace())

    assert calls == ["join", "integrate"]
    assert result.equals(integrated)


def test_build_report_returns_success_for_client_export(monkeypatch):
    """Report orchestration returns success when the client export completes."""
    empty = pd.DataFrame()
    raw = RawResult(empty, empty, empty)
    transformed = TransformedResult(empty, empty, empty, empty, empty)
    report = SimpleNamespace()
    calls = []

    monkeypatch.setattr(orchestrator, "load_data", lambda date: (empty, raw))
    monkeypatch.setattr(
        orchestrator, "transform_data", lambda raw_result, log, date: transformed
    )
    monkeypatch.setattr(orchestrator, "consolidate_info", lambda result: empty)
    monkeypatch.setattr(orchestrator, "get_summary", lambda *frames: report)
    monkeypatch.setattr(orchestrator, "get_differences", lambda frame, result: report)
    monkeypatch.setattr(orchestrator, "reload_differences", lambda *args: report)
    monkeypatch.setattr(orchestrator, "save_log", lambda frame, date: True)
    monkeypatch.setattr(
        orchestrator,
        "export_to_client_format",
        lambda *args: calls.append("client") or True,
    )

    result = orchestrator.build_report("2026_01", "cliente")

    assert result is True
    assert calls == ["client"]


def test_parse_args_reads_date_and_format(monkeypatch):
    """CLI parsing returns the requested date and output format."""
    monkeypatch.setattr(
        sys, "argv", ["report_pipeline", "--date", "2026_04", "--format", "winba"]
    )

    result = report_pipeline.parse_args()

    assert result.date == "2026_04"
    assert result.format == "winba"


def test_validate_date_format_accepts_valid_period():
    """Valid periods are returned unchanged."""
    assert report_pipeline.validate_date_format("2026_08") == "2026_08"


def test_validate_date_format_rejects_invalid_period():
    """Malformed periods raise the documented ValueError."""
    with pytest.raises(ValueError, match="YYYY_MM"):
        report_pipeline.validate_date_format("2026-08")


def test_normalize_concepts_creates_one_long_row():
    """Wide concept columns become normalized concept fields."""
    source = pd.DataFrame(
        {
            "UUID": ["u1"],
            "CONCEPTO1": ["renta"],
            "TOTALCONCEPTO1": [100],
            "CLAVEPRODSERVCONCEPTO1": ["0101"],
        }
    )

    result = normalize_concepts(source)

    assert result.loc[0, "CONCEPTO"] == "renta"
    assert result.loc[0, "TOTAL CONCEPTO"] == 100


def test_filter_metadata_info_keeps_only_prior_invoices_and_income_effect():
    """Metadata filtering removes other effects and future months."""
    source = pd.DataFrame(
        {
            "EfectoComprobante": ["I", "I", "E"],
            "FechaEmision": ["2026-01-05", "2026-03-05", "2026-01-06"],
            "Uuid": ["u1", "u2", "u3"],
        }
    )

    result = filter_metadata_info(source, "2026_02")

    assert result["Uuid"].tolist() == ["u1"]


def test_transform_cfdi_info_adds_usage_and_period_labels():
    """CFDI transformation maps usage and month labels for valid rows."""
    source = pd.DataFrame(
        {
            "TIPO_COMPROBANTE": ["I", "E"],
            "FECHA": ["2026-01-10", "2026-01-10"],
            "CFDI_USE": ["G03", "G03"],
            "CONCEPTO_IVA": [16, 8],
        }
    )

    result = transform_cfdi_info(source, "2026_01")

    assert result.shape[0] == 1
    assert result.loc[0, "USO CFDI"] == "Gastos en general"
    assert result.loc[0, "FECHA_PERIODO"] == "Enero"
